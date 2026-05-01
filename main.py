import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# 한글 폰트 설정 (macOS용 AppleGothic)
plt.rcParams['font.family'] = 'AppleGothic'
plt.rcParams['axes.unicode_minus'] = False

def save_visualizations(df, output_dir):
    """각 컬럼별로 개별 시각화 차트 생성 및 저장"""
    sns.set_theme(style="whitegrid", font='AppleGothic')
    viz_files = []

    # 1. 범주형 데이터 개별 시각화 (상위 10개 카테고리 빈도수)
    # product_category_name 등 주요 범주형 컬럼 대상
    cat_cols = df.select_dtypes(include=['object']).columns
    for col in cat_cols:
        if col == 'product_id': continue # ID는 시각화 제외
        
        plt.figure(figsize=(10, 6))
        counts = df[col].value_counts().head(15)
        sns.barplot(x=counts.values, y=counts.index, hue=counts.index, palette='husl', legend=False)
        plt.title(f'[{col}] 빈도수 분포 (Top 15)', fontsize=14)
        plt.xlabel('개수')
        plt.ylabel('항목')
        plt.tight_layout()
        
        file_path = os.path.join(output_dir, f'viz_cat_{col}.png')
        plt.savefig(file_path, dpi=200)
        plt.close()
        viz_files.append(('categorical', col, file_path))
        print(f"📊 범주형 차트 생성: {file_path}")

    # 2. 수치형 데이터 개별 시각화 (데이터 분포/히스토그램 형태의 바 차트)
    num_cols = df.select_dtypes(include=[np.number]).columns
    for col in num_cols:
        plt.figure(figsize=(10, 6))
        # 수치형 데이터는 구간(bin)을 나누어 빈도 시각화
        sns.histplot(df[col], kde=True, color='skyblue', bins=30)
        plt.title(f'[{col}] 데이터 분포', fontsize=14)
        plt.xlabel('값')
        plt.ylabel('빈도')
        plt.tight_layout()
        
        file_path = os.path.join(output_dir, f'viz_num_{col}.png')
        plt.savefig(file_path, dpi=200)
        plt.close()
        viz_files.append(('numerical', col, file_path))
        print(f"📊 수치형 차트 생성: {file_path}")
    
    return viz_files

def run_analysis_to_excel():
    # 경로 설정
    file_path = 'data/products_dataset.csv'
    img_path = 'images/image_668a63.png'
    output_dir = 'output'
    excel_path = f'{output_dir}/product_analysis_report.xlsx'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    try:
        # 1. 데이터 로드
        df = pd.read_csv(file_path)
        print("✅ 데이터를 성공적으로 불러왔습니다.")

        # 개별 시각화 실행
        viz_list = save_visualizations(df, output_dir)

        # 2. 통계 데이터 생성
        stats = df.describe().T
        stats = stats[['count', 'mean', 'std', 'min', '25%', '50%', '75%', 'max']]
        category_avg = df.groupby('product_category_name').mean(numeric_only=True)
        info_df = pd.DataFrame({
            '컬럼명': df.columns,
            '데이터 타입': df.dtypes.values,
            '결측치 아닌 개수': df.count().values,
            'Null 존재 여부': df.isnull().any().values
        })

        # 3. 엑셀 파일 저장
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            info_df.to_excel(writer, sheet_name='1.데이터_정보', index=False)
            stats.to_excel(writer, sheet_name='2.수치형_통계_상세')
            category_avg.to_excel(writer, sheet_name='3.카테고리별_평균')
            df.head(100).to_excel(writer, sheet_name='4.데이터_샘플(100개)', index=False)

        # 4. 엑셀에 개별 차트 삽입
        wb = load_workbook(excel_path)
        ws_viz = wb.create_sheet('5.시각화_상세분석')
        
        current_row = 2
        for v_type, col_name, f_path in viz_list:
            # 설명 텍스트 추가
            ws_viz.cell(row=current_row, column=2, value=f"📊 {col_name} 분석 차트 ({v_type})")
            
            # 이미지 삽입
            img = Image(f_path)
            img.width = 500
            img.height = 300
            ws_viz.add_image(img, f'B{current_row + 1}')
            
            # 다음 이미지를 위한 행 간격 띄우기
            current_row += 18 
        
        wb.save(excel_path)
        print(f"📸 모든 개별 차트({len(viz_list)}개)가 엑셀에 삽입되었습니다.")
        print(f"🎉 리포트 생성이 완료되었습니다: {excel_path}")

    except FileNotFoundError:
        print(f"❌ 에러: 파일을 찾을 수 없습니다.")
    except Exception as e:
        print(f"❌ 에러 발생: {e}")

if __name__ == "__main__":
    run_analysis_to_excel()